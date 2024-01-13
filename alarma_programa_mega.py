import time
import serial.tools.list_ports
import serial
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import simpledialog
from tkinter import Scrollbar, Text, Frame
from tkinter import scrolledtext
import datetime
import openpyxl
from datetime import datetime
import easygui
from tkinter import messagebox
import sys
from openpyxl.styles import numbers
import os
import re
nueva_etiqueta = None  # Declarar la variable globalmente
arduino = None  # Declarar la variable globalmente
valor_1 = 0  # Variable para el valor a enviar
dias_semana = ["Lunes      ", "Martes    ", "Miercoles", "Jueves     ", "Viernes    ", "Sabado    ", "Domingo"]
dias_indices = {0: "monday", 1: "tuesday", 2: "wednesday", 3: "thursday", 4: "friday", 5: "saturday", 6: "sunday"}
lista_dias =[]
lista_seleccionados=[]
nueva_hora=0
dia_actual = datetime.now().strftime("%A").lower()

# Verificar si el archivo Excel existe para habilitar los checkbuttons
archivo_excel_existe = os.path.isfile('datos.xlsx')
texto_boton = "Base de Datos Creada" if archivo_excel_existe else "Crear Base de Datos"



# Crear la función para ocultar el botón "boton_guardar"
def ocultar_boton_guardar():
    if archivo_excel_existe:
        boton_guardar.grid_forget()


def hacer_descripciones_solo_lectura():
    for entry in descripciones_alarmas:
        entry.config(state="readonly")
def hacer_hora_alarma_solo_lectura():
    for entry in entries_alarmas:
        entry.config(state="readonly")
                
        
def ingreso_datos_hora(event):
        #for entry in entries_alarmas:
    entry = event.widget  # Obtener el cuadro de texto que disparó el evento
    nombre = entry.grid_info()["row"]  # Obtener el número de fila del cuadro de texto
    valor_texto = entry.get()  # Obtener el valor de texto ingresado en el cuadro de texto

    primer_texto= "La Hora Actual es "
    segundo_texto = "Ingrese la nueva Hora con formato HH:MM:SS"
     
    hora_alarma = simpledialog.askstring("Hora Alarma {} ".format(nombre), "{}{}\n {}".format(primer_texto,valor_texto, segundo_texto))

    if hora_alarma is None:
        hora_alarma = "vacio"
        messagebox.showerror("Advertencia", "El formato de hora ingresado es inválido. Use el formato HH:MM:SS.")
    
     # Validar el formato de la hora ingresada
    formato_valido = re.match(r'^([01]\d|2[0-3]):([0-5]\d):([0-5]\d)$', hora_alarma)
   
         
    print("Cuadro de texto:", nombre)
    print("Valor de texto:", valor_texto)
    print("Nueva Hora" , hora_alarma)
    
        
    entry.config(state="normal")
    entry.delete(0, "end")  # Eliminar contenido existente en el cuadro de texto
    entry.insert("end", hora_alarma)  # Insertar el nuevo valor de hora_alarma
    entry.config(state="readonly")  # Establecer el cuadro de texto como solo lectura
    guardar_alarmas()

def ingreso_datos_descripcion(event):
    entry = event.widget  # Obtener el cuadro de texto que disparó el evento
    nombre = entry.grid_info()["row"]  # Obtener el número de fila del cuadro de texto
    valor_texto = entry.get()  # Obtener el valor de texto ingresado en el cuadro de texto

    descripcion = simpledialog.askstring("Descripción {}".format(nombre), "Ingrese la nueva descripción para la alarma {}".format(nombre))

    if descripcion is None:
        descripcion = "vacio"
        messagebox.showerror("Advertencia", "No se ingreso ninguna Descripcion.")
        
    print("Cuadro de texto:", nombre)
    print("Valor de texto:", valor_texto)
    print("Nueva descripción:", descripcion)

    entry.config(state="normal")
    entry.delete(0, "end")  # Eliminar contenido existente en el cuadro de texto
    entry.insert("end", descripcion)  # Insertar la nueva descripción
    entry.config(state="readonly")  # Establecer el cuadro de texto como solo lectura
    guardar_alarmas()

    
           
# testea que el CH340 este conectado al pc
def detectar_puerto_inicio():
    puertos_disponibles = serial.tools.list_ports.comports()
    for puerto in puertos_disponibles:
        if "Arduino Mega 2560" in puerto.description:
            return puerto.device
    # Mostrar mensaje en pantalla cuando no se encuentra el dispositivo
    messagebox.showinfo("Mensaje", "El dispositivo Arduino Mega 2560 no está conectado, conecte el USB ")
    sys.exit()
    return None

     
detectar_puerto_inicio()

def guardar_descripciones():
    try:
        wb = openpyxl.load_workbook('datos.xlsx')
        ws = wb.active
        
        for i, entry in enumerate(descripciones_alarmas):
            descripcion = entry.get()
            if descripcion == "":
                descripcion = "vacio"
            ws.cell(row=i+2, column=3, value=descripcion)  # Escribir el valor en la columna 'Descripción'
            
        
        wb.save('datos.xlsx')
    except FileNotFoundError:
        
        pass
    

def cargar_datos_puerto():
    try:
        wb = openpyxl.load_workbook('datos.xlsx')
        ws = wb.active
        valor_predeterminado = ws['D1'].value
        combo_puertos.set(valor_predeterminado)
    
    except FileNotFoundError:
        pass


   
    
    
def conectar_puerto():
    global arduino
   
        
    if arduino and arduino.is_open:
        # Si hay una conexión en curso, cerrarla y mostrar un mensaje
        arduino.close()
        print("Conexión interrumpida")
        boton_conectar.configure(text="Apagado", style="BotonNormal.TButton")
        boton_conectar.after(1000, lambda: boton_conectar.configure(state=tk.NORMAL))
        boton_conectar.configure(state=tk.DISABLED)
    else:
        # Obtener el puerto seleccionado en el ComboBox
        puerto_seleccionado = combo_puertos.get()

        # Validar si se seleccionó un puerto y el dispositivo CH340 está presente en él
        if puerto_seleccionado and validar_dispositivo_ch340(puerto_seleccionado):
            try:
                # Establecer la conexión serial con el puerto seleccionado
                arduino = serial.Serial(puerto_seleccionado, 9600)
                print("Conexión establecida en el puerto", puerto_seleccionado)
                boton_conectar.configure(text="Encendido", style="BotonConectado.TButton")
            except serial.SerialException:
                print("No se pudo establecer la conexión en el puerto", puerto_seleccionado)
        else:
            easygui.msgbox("No se encontró el dispositivo Arduino Mega 2560 en el puerto seleccionado. \n seleccione un puerto valido y oprima el boton que figura Apagado,\n sino se establece conxecion la alarma no funcionara. Si es la \n primera vez que enciende el programa ingrese los valores que desee \n en la solapa Alarma y guarde los cambios para que se genere \n la base de datos")
    if arduino and arduino.is_open:    

        nombres = []
        for entry in entries_alarmas:
            nombre = entry.get()  # Obtener el valor actual del cuadro de texto
            nombres.append(nombre)
            hora_actual = datetime.now().strftime("%H:%M:%S")

        x=0     
        if x != hora_actual:
            print(nombres[0]+" testeo 1")
            print(nombres[1]+" testeo 2")
            print(nombres[2]+" testeo 3")
            print(nombres[3]+" testeo 4")
            print(nombres[4]+" testeo 5")
            print(nombres[5]+" testeo 6")
            print(nombres[6]+" testeo 7")
            print(nombres[7]+" testeo 8")
            print(nombres[8]+" testeo 9")
            print(nombres[9]+" testeo 10")
    
            print (hora_actual + " Hora Actual")

            if  nombres[0]== hora_actual or nombres[1] == hora_actual or nombres[2] == hora_actual or nombres[3] == hora_actual or nombres[4] == hora_actual or nombres[5] == hora_actual or nombres[6] == hora_actual or nombres[7] == hora_actual or nombres[8] == hora_actual or nombres[9] == hora_actual :
                if dia_actual in lista_seleccionados:
                    print("ring")
                    print(dia_actual)
                    arduino.write(str(valor_1).encode())
            ventana.after(1000,testear_alarmas)  # Programar la próxima verificación después de 1000 milisegundo
           
            

def validar_dispositivo_ch340(puerto):
    puertos_disponibles = serial.tools.list_ports.comports()
    for puerto_disponible in puertos_disponibles:
        if puerto_disponible.device == puerto and "Arduino Mega 2560" in puerto_disponible.description:
            return True
    return False
    
# Buscar el dispositivo CH340 en los puertos COM disponibles
def detectar_puerto_ch340():
    puertos_disponibles = serial.tools.list_ports.comports()
    for puerto in puertos_disponibles:
        if "Arduino Mega 2560" in puerto.description:
            return puerto.device
    return None
# Obtener el puerto del dispositivo CH340
puerto_ch340 = detectar_puerto_ch340()
#if puerto_ch340:
#     print("El dispositivo se detecto en", puerto_ch340)
#else:
#     print("No se encontró el dispositivo CH340.")
    #***************************************************

def seleccionar_puerto(event):
    
    guardar_puerto_excel()
    
def validar_hora(hora):
    try:
        horas, minutos, segundos = hora.split(":")
        horas = int(horas)
        minutos = int(minutos)
        segundos = int(segundos)
        if 0 <= horas <= 23 and 0 <= minutos <= 59 and 0 <= segundos <= 59:
            return True
        else:
            return False
    except ValueError:
        return False

def guardar_alarmas():
    for i, entry in enumerate(entries_alarmas):
        hora = entry.get()
        if validar_hora(hora):
            alarmas[i] = hora
        else:
            entry.delete(0, tk.END)
            entry.insert(0, alarmas[i])
    

            
    guardar_datos()
    guardar_puerto_excel()
    guardar_descripciones() 
     # Agregar una columna al archivo Excel ----------------------agrega -----------------
    try:
        wb = openpyxl.load_workbook('datos.xlsx')
        ws = wb.active
        ws.insert_cols(5)  # Insertar una columna en la posición 4 (después de 'Descripcion')
        ws.cell(row=1, column=5, value='Corte')  # Establecer el encabezado de la nueva columna de horarios de corte
        ws.cell(row=1, column=6, value='Dias')  # Establecer el encabezado de la nueva columna de horarios de corte
        # Establecer el formato de hora "00:00:00" en las celdas de la columna "Corte"
        formato_hora = numbers.FORMAT_DATE_TIME3
        for fila in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for celda in fila:
                celda.number_format = formato_hora
                celda.value = "vacio"
        
        wb.save('datos.xlsx')
    except FileNotFoundError:
        pass
        

    # Guardar los cambios en el archivo Excel
    wb.save('datos.xlsx')
    escribir_dias_en_excel()

   
    


    for i in range(7):
        entry_dia = lista_dias[i][0]  # Obtener el Checkbutton de la lista existente
        dia_var = lista_dias[i][1]  # Obtener la variable asociada al Checkbutton

        dia_var.trace("w", on_checkbutton_change)  # Asociar la función de callback a la variable dia_var
        entry_dia.configure(state=tk.NORMAL)  # Habilitar el checkbutton
        entry_dia.grid(row=i + 2, column=10, padx=0)
       

    boton_guardar.config(text="Base de Datos Creada")
    #////////////////////////////////// carga de datos //////////////////////////////////
   
      
    

    
def guardar_datos():
    
       
    data = {'Alarma': range(1, 11), 'Hora': alarmas}
    df = pd.DataFrame(data)
    
    
    # Guardar el DataFrame en un archivo Excel
    df.to_excel('datos.xlsx', index=False)
    
    # Agregar una columna al archivo Excel solo si no existe
    if not os.path.isfile('datos.xlsx'):
        try:
            wb = openpyxl.load_workbook('datos.xlsx')
            ws = wb.active
            ws.insert_cols(5)  # Insertar una columna en la posición 6 (después de 'Hora')
            ws.cell(row=1, column=5, value='Dias')  # Establecer el encabezado de la nueva columna de días
            
            descripcion = "vacio"  # Crea valores vacíos para la columna de los días
            for row in range(2, 12):
                ws.cell(row=row, column=5, value="Dias")

         
            wb.save('datos.xlsx')
        except FileNotFoundError:
            pass
        
    print(df.columns)



def cargar_datos():
    try:
        df = pd.read_excel('datos.xlsx')
        valor_puerto = df['Descripcion'][0]  # Obtener el valor de la celda D1
                
        for i, entry in enumerate(entries_alarmas):
            entry.delete(0, tk.END)
            entry.insert(0, df['Hora'][i])
            entry.config(state="readonly")
            
        for i, entry in enumerate(descripciones_alarmas):
            entry.delete(0, tk.END)
            entry.insert(0, df['Descripcion'][i])  # Cargar los datos de las descripciones
            entry.config(state="readonly")

       # Insertar el segmento de código original aquí +++++++++++++++++++++++++++++++++++++++++
        lista_seleccionados = df['Dias'].tolist()
        for i, (entry_dia, dia_var) in enumerate(lista_dias):
           if lista_seleccionados[i] == 'vacio':
               dia_var.set(False)
           else:
               dia_var.set(True)

                #***********************************************************************************   
            
    except FileNotFoundError:
        # Ejecutar código cuando no se encuentra el archivo Excel
        for entry in descripciones_alarmas:
            entry.delete(0, tk.END)
            entry.insert(0, "vacio")  
        pass
    
def guardar_puerto_excel():
    global nueva_etiqueta  # Acceder a la variable global

    puerto = combo_puertos.get()

    # Leer el archivo Excel existente
    try:
        wb = openpyxl.load_workbook('datos.xlsx')
        ws = wb.active
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo con encabezados
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Alarma', 'Hora', 'Descripcion', 'Dias'])
         # Agregar una columna para guardar los datos de descripción
        ws.insert_cols(3)  # Insertar una columna en la posición 3 (después de 'Hora')
        ws.cell(row=1, column=3, value='Descripcion')  # Establecer el encabezado de la columna
       
       
    # Actualizar el valor del puerto
    ws['C1'] = 'Descripcion'
    ws['D1'] = puerto

    # Guardar los cambios en el archivo Excel
    wb.save('datos.xlsx')

    try:
        wb = openpyxl.load_workbook('datos.xlsx')
        ws = wb.active
        valor_celda = "El puerto guardado es:" + ws['D1'].value
        valor_com.set(valor_celda)
        
        if ws['D1'].value == puerto_ch340:
            if nueva_etiqueta is None:
                # Crear y mostrar la nueva etiqueta
                nueva_etiqueta = ttk.Label(frame_puertos, text="Disp.detectado en: " + puerto_ch340, foreground="red")
                nueva_etiqueta.pack()
            else:
                # Actualizar el texto de la etiqueta existente
                nueva_etiqueta.config(text="Disp.detectado en: " + puerto_ch340, foreground="red")
                
        if ws['D1'].value != puerto_ch340:
            if nueva_etiqueta is None:
                # Crear y mostrar la nueva etiqueta
                nueva_etiqueta = ttk.Label(frame_puertos, text="No hay dispositivos en este puerto \n utilice el puerto: " + puerto_ch340)
                nueva_etiqueta.pack()
            else:
                # Actualizar el texto de la etiqueta existente
                nueva_etiqueta.config(text="No hay dispositivos en este puerto \n utilice el puerto: " + puerto_ch340)
                
    except FileNotFoundError:
        valor_com.set("No hay puertos")  # Si el archivo no existe, establecer el valor en blanco



        
def obtener_valor_celda():
    global nueva_etiqueta  # Acceder a la variable global

    try:
        wb = openpyxl.load_workbook('datos.xlsx')
        ws = wb.active
        valor_celda ="El puerto guardado es:" + ws['D1'].value
        valor_com.set(valor_celda)

        if ws['D1'].value == puerto_ch340:
           # Crear y mostrar la nueva etiqueta
            frame_puertos.pack(fill="both", expand=True)

            nueva_etiqueta = ttk.Label(frame_puertos, text="Disp.detectado en: " +puerto_ch340, foreground="red")
                       
            valor_celda ="El puerto guardado es:" + ws['D1'].value
            valor_com.set(valor_celda)
            nueva_etiqueta.pack()
        if ws['D1'].value != puerto_ch340:
           # Crear y mostrar la nueva etiqueta
            nueva_etiqueta = ttk.Label(frame_puertos, text="No hay dispositivos en este puerto \n actualize su ubicacion a:" + puerto_ch340) 
            nueva_etiqueta.pack()
            
            
    except FileNotFoundError:
        valor_com.set("No hay puertos")  # Si el archivo no existe, establecer el valor en blanco

        

#*************************************************** base de datos************
def ver_datos():    
    
     # Ruta del archivo Excel en el escritorio
    ruta_archivo = "~/Desktop/datos.xlsx"
     
    # Crear ventana principal
    ventana = tk.Tk()
    ventana.title("Datos del archivo Excel")

    # Crear pestaña
    pestaña = ttk.Frame(ventana)
    pestaña.pack()

    # Cargar el archivo Excel en un DataFrame
    df = pd.read_excel(ruta_archivo)

    # Crear tabla para mostrar los datos del DataFrame
    tabla = ttk.Treeview(pestaña)
    tabla.pack()

    # Agregar encabezados de columnas
    encabezados = df.columns.tolist()
    tabla['columns'] = encabezados
    for columna in encabezados:
        tabla.heading(columna, text=columna)

    # Agregar filas de datos
    datos = df.to_numpy().tolist()
    for fila in datos:
        tabla.insert('', tk.END, values=fila)
#//////////////////////////////////////////////////////////////////////
def testear_alarmas():
    if arduino and arduino.is_open:    

        nombres = []
        for entry in entries_alarmas:
            nombre = entry.get()  # Obtener el valor actual del cuadro de texto
            nombres.append(nombre)
            print(nombre)
            hora_actual = datetime.now().strftime("%H:%M:%S")

        x=0     
        if x != hora_actual:
            print(nombres[0]+" alarma 1")
            print(nombres[1]+" alarma 2")
            print(nombres[2]+" alarma 3")
            print(nombres[3]+" alarma 4")
            print(nombres[4]+" alarma 5")
            print(nombres[5]+" alarma 6")
            print(nombres[6]+" alarma 7")
            print(nombres[7]+" alarma 8")
            print(nombres[8]+" alarma 9")
            print(nombres[9]+" alarma 10")
    
            print (hora_actual + " Hora Actual")
            
            if  nombres[0]== hora_actual or nombres[1] == hora_actual or nombres[2] == hora_actual or nombres[3] == hora_actual or nombres[4] == hora_actual or nombres[5] == hora_actual or nombres[6] == hora_actual or nombres[7] == hora_actual or nombres[8] == hora_actual or nombres[9] == hora_actual :
                if dia_actual in lista_seleccionados:
                    print("ring")
                    print(dia_actual)
                    arduino.write(str(valor_1).encode())
            ventana.after(1000,testear_alarmas)  # Programar la próxima verificación después de 500 milisegundo
   

  #****************************fin base de datos ********************************  
  

# Crear ventana principal

ventana = tk.Tk()
ventana.geometry("500x420")
ventana.title("Alarma")

# Crear estilo para el botón en estado normal
ventana.style = ttk.Style()
ventana.style.configure("BotonNormal.TButton", foreground="black", background="red", font=("Arial", 16))  # Aumentar el tamaño de la fuente


# Crear estilo para el botón en estado conectado
ventana.style.configure("BotonConectado.TButton", foreground="black", background="aquamarine1", font=("Arial", 16))

# Crear estilo para el botón de guardar en la pestaña 2
ventana.style.configure("BotonGuardar.TButton", foreground="black", background="light blue", font=("Arial", 12))

# Crear estilo para el botón de guardar en la pestaña 2
ventana.style.configure("boton_ver_datos", foreground="black", background="light blue", font=("Arial", 12))

# Variables
valor_com = tk.StringVar()
alarmas = ["vacio"] * 10



# Crear pestañas
pestañas = ttk.Notebook(ventana)
pestaña1 = ttk.Frame(pestañas)
pestaña2 = ttk.Frame(pestañas)
pestaña3 = ttk.Frame(pestañas)
pestañas.add(pestaña2, text="Alarmas")
pestañas.add(pestaña1, text="Configuracion")
pestañas.add(pestaña3, text="Ayuda")
pestañas.pack(fill="both", expand=True)
pestañas.configure(width=200, height=200)

# Solapa de puertos
frame_puertos = ttk.Frame(pestaña1)
frame_puertos.pack()

label_puertos = ttk.Label(frame_puertos, text="Seleccione un puerto:")
label_puertos.pack()
#******************************* lista de  puertos disponibles *********************************
# Obtener una lista de los puertos COM disponibles
puertos_disponibles = serial.tools.list_ports.comports()

# Almacenar los puertos COM disponibles en una lista
values = []
for puerto in puertos_disponibles:
    values.append(puerto.device)

combo_puertos = ttk.Combobox(frame_puertos, values=values)

#********************************************************************************************
combo_puertos = ttk.Combobox(frame_puertos, values=values)
combo_puertos.current(0)  # Establecer el índice de la opción seleccionada (0 para "con0")
combo_puertos.bind("<<ComboboxSelected>>", seleccionar_puerto)
combo_puertos.pack()

# Cargar los datos predeterminados del puerto
cargar_datos_puerto()

label_valor_com = ttk.Label(frame_puertos, textvariable=valor_com)
label_valor_com.pack()

# Solapa de alarmas
frame_alarmas = ttk.Frame(pestaña2)
frame_alarmas.pack()

# Solapa de Movimientos
frame_Movimientos = ttk.Frame(pestaña3)
frame_Movimientos.pack()


# establece las los labeltext y los textbox de las solapas ///////////////////////////////////////////////
entries_alarmas = []
descripciones_alarmas = []


for i in range(10):
    label_alarma = ttk.Label(frame_alarmas, text="Alarma {}: ".format(i + 1))
    label_alarma.grid(row=i + 2, column=0, sticky="e")
    entry_alarma = ttk.Entry(frame_alarmas, width=10)
    entry_alarma.insert(0, alarmas[i])
    entry_alarma.grid(row=i + 2, column=1)
    entries_alarmas.append(entry_alarma)
    entry_alarma.bind("<Button-1>",ingreso_datos_hora)  # Asignar la función al evento <FocusIn> de cada cuadro de texto

    
    label_alarma.config(font=("TkDefaultFont", 10, "bold"))


    descripcion_alarma = ttk.Label(frame_alarmas, text="")
    descripcion_alarma.grid(row=i + 2, column=2, sticky="e", padx=2)
    entry_descripcion = ttk.Entry(frame_alarmas, width=20)
    entry_descripcion.grid(row=i + 2, column=3)
    descripciones_alarmas.append(entry_descripcion)
    entry_descripcion.bind("<Button-1>", ingreso_datos_descripcion)


dia_etiqueta = ttk.Label(frame_alarmas, text="Días        ")
dia_etiqueta.grid(row=i-8, column=9, sticky="e")
dia_etiqueta.config(font=("TkDefaultFont", 10, "bold"))

descripcion_etiqueta = ttk.Label(frame_alarmas, text=" Descripcion             ")
descripcion_etiqueta.grid(row=i-8, column=3, sticky="e", padx=0)
descripcion_etiqueta.config(font=("TkDefaultFont", 10, "bold"))

descripcion_etiqueta = ttk.Label(frame_alarmas, text=" Horario        ")
descripcion_etiqueta.grid(row=i-8, column=1, sticky="e", padx=0)
descripcion_etiqueta.config(font=("TkDefaultFont", 10, "bold"))


#******************checkbuttons *********************************************************************
dias_semana = ["Lunes      ", "Martes    ", "Miercoles", "Jueves     ", "Viernes    ", "Sabado    ", "Domingo"]
dias_indices = {0: "monday", 1: "tuesday", 2: "wednesday", 3: "thursday", 4: "friday", 5: "saturday", 6: "sunday"}
lista_dias =[]




def imprimir_checkbuttons_seleccionados():
    checkbuttons_seleccionados = []
    for i, (entry_dia, dia_var) in enumerate(lista_dias):
        if dia_var.get():
            checkbuttons_seleccionados.append(dias_indices[i])
    return checkbuttons_seleccionados



def escribir_dias_en_excel():

    wb = openpyxl.load_workbook('datos.xlsx')
    ws = wb.active
    
    if dias_indices[0] in lista_seleccionados:
        ws.cell(row=2, column=6, value="lunes")
    else:     
        ws.cell(row=2, column=6, value="vacio")
    

    if dias_indices[1] in lista_seleccionados:
        ws.cell(row=3, column=6, value="martes")
    else:     
        ws.cell(row=3, column=6, value="vacio")

    if dias_indices[2] in lista_seleccionados:
        ws.cell(row=4, column=6, value="miercoles")
    else:     
        ws.cell(row=4, column=6, value="vacio")

    if dias_indices[3] in lista_seleccionados:
        ws.cell(row=5, column=6, value="jueves")
    else:     
        ws.cell(row=5, column=6, value="vacio")

    if dias_indices[4] in lista_seleccionados:
        ws.cell(row=6, column=6, value="viernes")
    else:     
        ws.cell(row=6, column=6, value="vacio")

    if dias_indices[5] in lista_seleccionados:
        ws.cell(row=7, column=6, value="sabado")
        wb.save('datos.xlsx')
    else:     
        ws.cell(row=7, column=6, value="vacio")
        wb.save('datos.xlsx')  

    if dias_indices[6] in lista_seleccionados:
        ws.cell(row=8, column=6, value="domingo")
        wb.save('datos.xlsx')
    else:     
        ws.cell(row=8, column=6, value="vacio")
        wb.save('datos.xlsx')



def on_checkbutton_change(*args):
    checkbuttons_seleccionados = imprimir_checkbuttons_seleccionados()
    lista_seleccionados.clear()
    lista_seleccionados.extend(checkbuttons_seleccionados)
    #print(lista_seleccionados , "de checkbutton_change")
    #print(dias_indices[0], "de checkbutton_change")

    
    escribir_dias_en_excel()
    
   
    
for i in range(7):
    dia_activar = ttk.Label(frame_alarmas, text="{} ".format(dias_semana[i]))
    dia_activar.grid(row=i + 2, column=9, sticky="e", padx=15)
    dia_var = tk.BooleanVar()
    dia_var.trace("w", on_checkbutton_change)  # Asociar la función de callback a la variable dia_var
    entry_dia = ttk.Checkbutton(frame_alarmas, variable=dia_var, state=tk.DISABLED if not archivo_excel_existe else tk.NORMAL)# checkbutton desabilitados
    entry_dia.grid(row=i + 2, column=10, padx=0)
    lista_dias.append((entry_dia, dia_var))
    

    
    
# ******************fin checkbuttons************************
    
cargar_datos()
obtener_valor_celda()
testear_alarmas()

hacer_descripciones_solo_lectura()
hacer_hora_alarma_solo_lectura()

#espacio en blanco para separar titulos 
texto_alarmas = ttk.Label(frame_alarmas, text=" ")
texto_alarmas.grid(row=1, column=0,pady=40,)

# Botón de guardar alarmas
boton_guardar = ttk.Button(frame_puertos, text=texto_boton, style="BotonGuardar.TButton", command=guardar_alarmas)
boton_guardar.pack(pady=10)

boton_ver_datos = ttk.Button(frame_puertos, text="Ver Base de Datos", style="BotonGuardar.TButton", command=ver_datos)
boton_ver_datos.pack(pady=20)

boton_conectar = ttk.Button(ventana, text="Apagado", style="BotonNormal.TButton", command=conectar_puerto)
boton_conectar.configure(width=200)  # Ajustar el ancho del botón a 200 píxeles
boton_conectar.pack()

ocultar_boton_guardar()

# Crear un Frame para el contenedor
frame_Movimientos.pack(fill=tk.BOTH, expand=True)

# Crear una barra de desplazamiento vertical
scrollbar = Scrollbar(frame_Movimientos)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Crear un widget Text dentro del Frame
text_widget = Text(frame_Movimientos, yscrollcommand=scrollbar.set, spacing1=0, spacing2=0)
text_widget.pack()

# Configurar la barra de desplazamiento para interactuar con el widget Text
scrollbar.config(command=text_widget.yview)

# Insertar texto predeterminado
texto4="\n Ingreso de datos: \n Para insertar o cambiar un dato clickee con el mouse sobre la casilla"\
       " deseada, para seleccionar los dias en que las alarmas se repetiran basta con hacer un click"\
       " en el cuadrado correspondiente y aparecera con un tilde, haga otro clicck si quiere desmarcar"\
       " el dia.\n\n Configuracion: \n Si es la primera vez que abre el programa tendra que selecionar"\
       " un puerto para el dispositivo. El programa le mostrara donde se encuentra el Dispositivo "\
       "usted debe ir al menu desplegable y elegir el COM correspondiente a su dispositivo.\n\n"\
       "Preparar sus Datos: \n Al inicio del programa y por unica vez devera geberar la base de datos"\
       "apretando el boton Crear base de datos, esto genera un archivo excel en el cual se almacenaran"\
       "las horas que ingreso, los dias que eligio y las descripciones.\n"\
       "\n Atencion: \n El programa biene en dos versiones una corresponde a la placa CH340 y la otra a Arduino Mega 2560"\
       "en caso de reparacion recuerde que depende con que placa reemplazara y en consecuencia instalar el programa correspondiente"\
       "de no instalar el programa correspondiente el programa no abrira y le saldra un mensaje diciendo que el dispositivo no se encuentra"\
       " de la misma forma como si el dispositivo no estuviera conectado."\
       " \n\n Conflictos:\n El programa usa el puerto serial del pc para envia una señal al dispositivo, el progragra"\
       " CURA de impresion 3D es un programa que suele enviarseñales a los puertos serie, con lo cual la alarma sonara en repetidas ocaciones"\
       " al ejecutarlo, se recomienda desconectar el dispositivo o anular temporalmente el puerto serie antes de usarlo.\n"
       

text_widget.insert(tk.END, texto4)
MAX_LINES = 300  # Número máximo de líneas en el widget Text

def imprimir(texto):
    text_widget.insert(tk.END, texto + "\n")
    text_widget.see(tk.END)  # Desplazar al final del texto
    # Verificar la cantidad de líneas y eliminar las más antiguas si es necesario
    line_count = int(text_widget.index(tk.END).split(".")[0])
    if line_count > MAX_LINES:
        text_widget.delete("1.0", f"{line_count - MAX_LINES}.0")
    

# Redirigir la salida de print al widget de texto
#sys.stdout.write = imprimir


# Ejecutar ventana
ventana.mainloop()

#*************************************para primera casilla ******************


# Cierra la conexión serial

#arduino.close()
